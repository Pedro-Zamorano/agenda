# from openpyxl import Workbook
# import os.path
from openpyxl import load_workbook

filesheet = "./libro_registro.xlsx"

def func_apunte(ingreso1):
    # LOGRAR QUE LA DATA INGRESADA SE VAYA A LA HOJA "APUNTES" // LOGRADO! pero es mejorable
    wb = load_workbook(filesheet)
    hoja_A = wb.get_sheet_by_name('APUNTES')
    wb.active = hoja_A

    lista_previa = []
    lista_previa.append(ingreso1)
    sheet = wb.active

    for row in lista_previa:
        sheet.append(row)

    wb.save(filesheet)


def func_recordatorio(ingreso):
    wb = load_workbook(filesheet)
    hoja_R = wb.get_sheet_by_name('RECORDATORIO')
    wb.active = hoja_R

    lista_previa = []
    lista_previa.append(ingreso)
    sheet = wb.active

    for row in lista_previa:
        sheet.append(row)

    wb.save(filesheet)

"""
if not os.path.isfile(filesheet):
    wb = Workbook()
    print("El archivo NO existe")
    hoja_R = wb.create_sheet("RECORDATORIO")
    hoja_R.append(('FECHA', 'RECORDATORIO'))

    hoja_A = wb.create_sheet("APUNTES")
    hoja_A.append(('FECHA', 'APUNTE'))

    wb.save(filesheet)
"""