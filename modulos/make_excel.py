from openpyxl import Workbook
import os.path

filesheet = "./libro_registro.xlsx"


def check():
    if not os.path.isfile(filesheet):
        wb = Workbook()
        hoja_R = wb.create_sheet("RECORDATORIO")
        hoja_R.append(('FECHA', 'RECORDATORIO'))

        hoja_A = wb.create_sheet("APUNTES")
        hoja_A.append(('FECHA', 'APUNTE'))

        wb.save(filesheet)
